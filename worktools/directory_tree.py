#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2020/10/21 11:57
# @Author  : 李加林
# @FileName: directory_tree.py
# @Software: PyCharm
# @Blog    ：https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/

"""
从一个文件夹路径生成目录树（常用格式文件会被写入），并写入excel文档（带超链接）
"""

import os
import openpyxl
import PySimpleGUI as sg
from openpyxl.styles import Font
# from openpyxl.utils import get_column_letter

class DirectoryTree(object):
    """
    目录树生成并写入excel
    """
    def __init__(self, folder_path):
        """
        :param folder_path: 文件夹路径
        :type folder_path: str
        """
        self.folder_path = folder_path
        # 创建空列表用于存放文件夹下所有文件的绝对路径
        self.absolute_path_list = []
        # 自定义常见文件类型，文件夹里面文件的类型符合才会被写入excel
        self.filetype = ('.doc',
                         '.docx',
                         '.xls',
                         '.xlsx',
                         '.ppt',
                         '.pptx',
                         '.pdf',
                         '.jpg',
                         '.jpeg',
                         '.png',
                         '.bmp',
                         '.gif',
                         '.txt',
                         '.csv',
                         '.md',
                         '.py',
                         '.mp4',
                         '.avi',
                         '.mkv',
                         '.mp3',
        )
        # 计算文件夹路径长度，用于后续计算相对路径
        self.path_lenth = len(folder_path)
        # 执行

    def set_filetype(self, filetype: list or tuple):
        """
        设置需要的文件格式
        """
        self.filetype = filetype

    def get_absolute_path(self) :
        """
        获取文件夹下所有绝对路径文件名
        """
        for root, dirs, files in os.walk(self.folder_path):
            for file in files:
                if os.path.splitext(file)[1] in self.filetype:
                    self.absolute_path_list.append(os.path.join(root, file))

    def write_workbook(self):
        """
        按步骤将文件名拆分并写入excel，形成目录树
        """
        # 创建工作薄用于写目录树
        wb = openpyxl.Workbook()
        ws = wb.active
        # 设置字体类型（该字体类型为添加下划线）
        font = Font(underline='single')
        ws['A1'].value = self.folder_path+'\\'
        # 设置超链接
        ws['A1'].hyperlink = self.folder_path
        # 应用字体
        ws['A1'].font = font
        # row,column 是临时的变量，用于逐行或逐列按需写入数据用
        row = 2
        for each_path in self.absolute_path_list:
            # 计算相对路径，并按文件夹深度进行拆分
            relative_path = each_path[self.path_lenth+1:]
            r_list = relative_path.split('\\')
            column = 2
            for e in r_list:
                # 判断拆分后的路径是文件夹还是文件
                if os.path.splitext(e)[1] not in self.filetype :
                    ws.cell(column=column, row=row).value = e + '\\'
                else:
                    ws.cell(column=column, row=row).value = e
                    # 添加超链接与下划线
                    ws.cell(column=column, row=row).hyperlink = r'' + each_path
                    ws.cell(column=column, row=row).font = font
                column += 1
            row += 1
        wb.save("tree.xlsx")
        # 保存后直接打开excel文件
        # os.startfile("tree.xlsx")

    def creat_directorytree(self):
        self.get_absolute_path()
        self.write_workbook()

class DirectoryTreeGUI(object):

    def __init__(self):
        # 设置pysimplegui主题，不设置的话就用默认主题
        # sg.ChangeLookAndFeel('Purple')
        # 定义2个常量，供下面的layout直接调用，就不用一个个元素来调字体了
        # 字体和字体大小
        self.FONT = ("微软雅黑", 16)
        # 可视化界面上元素的大小
        self.SIZE = (20, 1)
        # 界面布局
        self.layout = [
            # sg.Image()插入图片，支持gif和png
            [sg.Image(filename="images/peppa.png",pad=(130,0))],
            # sg.Text()显示文本
            [sg.Text('', font=self.FONT, size=self.SIZE)],
            # sg.Input()是输入框
            [sg.Text('请选择要生成目录树的文件夹：', font=self.FONT, size=(30, 1))],
            [sg.Input(key="_FOLDER_", readonly=True, size=(36, 1), font=self.FONT),
             sg.FolderBrowse(button_text='选择文件夹', size=(10, 1), font=self.FONT)],
            [sg.Text('')],
            [sg.Btn('生成目录树', key='_TREE_', font=("微软雅黑", 16), size=(23, 1)),
             sg.Input(key='_RESULT_', readonly=True, size=(10, 1), font=self.FONT)],
        ]
        # 创建窗口，引入布局，并进行初始化
        # 创建时，必须要有一个名称，这个名称会显示在窗口上
        self.window = sg.Window('目录树生成工具by李加林', layout=self.layout, finalize=True)

        # 窗口持久化
    def run(self):
        # 创建一个事件循环，否则窗口运行一次就会被关闭
        while True:
            # 监控窗口情况
            event, value = self.window.Read()
            # 当获取到事件时，处理逻辑（按钮绑定事件，点击按钮即触发事件）
            if event == '_TREE_':
                folder = value['_FOLDER_']
                # 生成目录树
                tree = DirectoryTree(folder_path=folder)
                tree.creat_directorytree()
                # 函数完成后返回处理完成标志到窗口界面上
                self.window.Element("_RESULT_").Update("处理完成！")
            # 如果事件的值为 None，表示点击了右上角的关闭按钮，则会退出窗口循环
            if event is None:
                break
        self.window.close()

if __name__ == '__main__':
    dt = DirectoryTreeGUI()
    dt.run()
    # tree = DirectoryTree(r"D:\MyNutstore\工作文档\01_工作项目\03_党风廉政建设@[英德市纪委监委]\[长期工作]英德市主体责任和勤廉监督业务平台（按要求上报材料）")
    # tree.creat_directorytree()


