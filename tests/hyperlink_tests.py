#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2020/10/11 23:32
# @Author : 李加林
# @FileName : hyperlink_tests.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/

import openpyxl

# from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# wb = openpyxl.load_workbook("1.xlsx")
# ws =wb.active
# ws.cell(row=1,column=1,value=11111111)
# # ft = Font(underline=True)
# ws.cell(row=8,column=1).value = "超链接测试"
# ws.cell(row=8,column=1).hyperlink = r"D:\MyNutstore"
# # ws.cell(row=8,column=1).font = ft
# wb.save("done.xlsx")
p = r"转发关于进一步做好审计检察整改落实相关工作的通知"
r = r"D:\MyNutstore\工作文档\01_工作项目\01_社保基金安全@[广东省人社厅,清远市人社局,清远市社保局,英德市人社局]\20200927_[基于粤社保办〔2020〕58 号）]转发关于进一步做好审计检察整改落实相关工作的通知"
print(p in r)