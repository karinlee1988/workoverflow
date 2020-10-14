#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2020/10/12 16:49
# @Author  : 李加林
# @FileName: search.py
# @Software: PyCharm
# @Blog    ：https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/

"""
本模块用于批量查找搜索数据
"""
import pandas as pd
import openpyxl
from workoverflow.fileio import get_allfolder_fullfilename,record_csv

class ElementSearch(object):
    """
    用于某个数据(或excel表中某列数据)在一堆其他excel文件中查找,找出对应行后进行记录.
    """

    def __init__(self, source_path:str, element:str) -> None:
        """
        :param source_path: 要查找的文件夹路径(文件夹下包括子文件夹的相关文件都会搜索）
        :type  source_path: str

        :param element: 要搜索的关键字（需完全匹配）
        :type  element: str

        :return: None
        """
        self.source_path = source_path
        self.element = element

    @staticmethod
    def transfer(x) -> str:
        """
        将数据处理后再写入csv，避免长数字写入时出现科学计数法现象
        :param x: 待写入的数据
        :type  x: int or str

        :return :处理后的数据
        :rtype : str
        """
        return '\t' + str(x)

    def search_element(self):
        """
        使用openpyxl开始进行搜索
        """
        # 获取文件名，只支持xlsx
        file_list = get_allfolder_fullfilename(self.source_path, [".xlsx", ])
        # 初始化csv，用于存放搜索结果
        record_csv(['文件路径','工作表名','行号','列号'], self.element + "_搜索结果.csv")
        # 遍历文件名列表，取得Workbook对象和Worksheet对象，开始遍历查找
        for filename in file_list:
            wb = openpyxl.load_workbook(filename)
            for sheet in wb:
                for row in range(1, sheet.max_row + 1):
                    for column in range(1, sheet.max_column + 1):
                        if sheet.cell(row=row, column=column).value == self.element:
                            # row_list 先初始化，添加一些索引信息
                            row_list = [filename,sheet.title,row,column,]
                            for column_result in range(1,sheet.max_column+1):
                                # row_list 添加查找数据所在行的值（值统一转为str）。\t 是为了添加身份证等长数字后，
                                # 写入csv时不会出现科学计数法现象。
                                row_list.append('\t' + str(sheet.cell(row=row, column=column_result).value))
                            # 数据写入，保存为搜索结果
                            record_csv(row_list, self.element + "_搜索结果.csv")
                            break

    def search_element_plus(self):
        """
        升级版方式，使用pandas进行搜索，速度相对更快
        """
        # 获取文件名，支持xls，xlsx
        file_list = get_allfolder_fullfilename(self.source_path, [".xlsx",".xls" ])
        # 初始化csv，用于存放搜索结果
        record_csv(['文件路径', '工作表名', '行号', '列号'], self.element + "_搜索结果.csv")
        for filename in file_list:
            # pd.read_excel读取文件为dataframe
            # sheetsheet_name=None时返回一个以sheet名称（str）为key，对应sheet表格内容（dataframe）为value的字典
            df_dict = pd.read_excel(filename, header=None, sheet_name=None)
            # 遍历所有sheet页，通过字典获取对应sheet的dataframe
            for sheetname in df_dict.keys():
                df_sheet = df_dict[sheetname]
                # 转换成列表，enumerate()添加索引（从1开始）
                list_sheet = df_sheet.values.tolist()
                for row_index, row in enumerate(list_sheet, start=1):
                    for column_index, e in enumerate(row, start=1):
                        # print(row_index,column_index,e)
                        # 查找到对应值后，构建一维列表，写入csv
                        if e == self.element:
                            # 通过map函数，用自定义的transfer()对原始数据列表进行转换
                            list_content = list(map(self.transfer,list_sheet[row_index-1]))
                            row_list = [filename, sheetname, row_index,column_index] + list_content
                            record_csv(row_list,self.element + "_搜索结果.csv")
                            break


if __name__ == '__main__':
    s = ElementSearch(r"D:\MyNutstore\PersonalStudy\Python\WorkOverflow\tests\source","440228196608156369")
    s.search_element_plus()

