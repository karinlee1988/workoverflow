#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2020/10/6 14:16
# @Author : 李加林
# @FileName : exceler.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/

"""
本模块用于对excel表格进行各种处理操作
"""

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string  # ,get_column_letter
from pandas.core.frame import DataFrame
from workoverflow.fileio import *

def vlookup(
        wb_template:object,
        ws_template_index:int,
        template_key:str,
        template_value:str,
        wb_source:object,
        ws_source_index:int,
        source_key:str,
        source_value:str,
        line:int
        ) -> None:
    """
    对2个不同的工作薄执行vlookup操作

    模板工作薄：需要写入数据的工作薄
    数据工作薄：根据模板工作薄提供的条件（列）在数据工作薄中查找，提供数据来源的工作薄
    注意！函数执行完后，只对wb_template对象进行了数据写入。在函数外部还需wb_template.save("filename.xlsx"),vlookup后的数据才能保存为excel表。

    :param:
        'wb_template': 模板工作薄对象
        'ws_template_index': 需要处理的模板工作表索引号
        'template_key': 模板工作表key所在列号
        'template_value': 模板工作表value需要填写的列号
        'wb_source': 数据工作薄对象
        'ws_source_index': 需要处理的数据工作表索引号
        'source_key':  数据工作表key所在列号
        'source_value': 数据工作表value所在列号
        'line' :从第几行开始vlookup

    :type:
        'wb_template': class Workbook
        'ws_template_index': int
        'template_key': str
        'template_value': str
        'wb_source': class Workbook
        'ws_source_index': int
        'source_key':  str
        'source_value': str
        'line' :int

    :return: None

    """
    # 获取数据工作表对象
    ws_source = wb_source[wb_source.sheetnames[ws_source_index]]
    # 获取模板工作表对象
    ws_template = wb_template[wb_template.sheetnames[ws_template_index]]
    # 获取数据工作表的查找列和数据列，分别生成2个元组
    source_key_tuple = ws_source[source_key]
    source_value_tuple = ws_source[source_value]
    # 创建2个列表，遍历元组，将元组中每个单元格的值添加到列表中
    list_key =[]
    list_value = []
    for cell in source_key_tuple:
        list_key.append(cell.value)
    for cell in source_value_tuple:
        list_value.append(cell.value)
    # 通过数据工作表的key列和value列，创建好需要进行vlookup的字典
    dic = dict(zip(list_key,list_value))
    # 将列号 （str）转为列索引值 （int）
    template_key_index = column_index_from_string(template_key)
    template_value_index = column_index_from_string(template_value)
    # 从第line行开始进行vlookup  可根据表头行数进行修改
    for row in range(int(line),ws_template.max_row+1):
        # ------------------------------------------------------
        # 采用dict[key]方式查字典，如果没有key的话会raise KeyError
        # try:
        #     ws_template.cell(row=row,column=template_value_index).value = dic[
        #     ws_template.cell(row=row,column=template_key_index).value]
        # except KeyError:
        #     #找不到数据 相应的单元格填上#N/A
        #     ws_template.cell(row=row, column=template_value_index).value = "#N/A"
        # -------------------------------------------------------
        #采用dict.get()避免出现keyerror
        ws_template.cell(row=row, column=template_value_index).value = dic.get(
            ws_template.cell(row=row, column=template_key_index).value)

def worksheet_save_as(workbook:object) -> None:
    """
    将一个工作薄里面的多个工作表分别另存为独立的工作薄，独立的工作薄名称为原工作薄各工作表表名

    :param workbook:需要进行工作表另存为的workbook对象
    :type workbook: class Workbook

    :return: None

    """
    sheetname_list = workbook.sheetnames
    for name in sheetname_list:
        worksheet = workbook[name]
        # 创建新的Excel
        workbook_new = openpyxl.Workbook()
        # 获取当前sheet
        worksheet_new = workbook_new.active
        # 两个for循环遍历整个excel的单元格内容
        for i, row in enumerate(worksheet.iter_rows(),start=1): #enumerate()从1开始  # 或者for i, row in enumerate(worksheet.rows):
            for j, cell in enumerate(row,start=1):#enumerate()从1开始
                # 写入新Excel
                worksheet_new.cell(row=i, column=j, value=cell.value)
                # 设置新Sheet的名称
                worksheet_new.title = name
        workbook_new.save(name + '.xlsx')

def split_workbook_by_column(workbook_path:str,
                             sheet_index:int,
                             title_index:int,
                             column_index:int) -> None:
    """
    将一个excel工作薄中指定的工作表，按列字段进行拆分，拆分后的多个excel表格分别以列字段命名
    使用pandas库处理，支持.xlsx类型

    :param workbook_path: 要拆分的工作薄文件路径
    :type workbook_path:  str

    :param sheet_index: 要拆分的工作薄的工作表序号（从0开始）
    :type sheet_index:  int

    :param title_index: 拆分后表头的行数（从1开始）
    :type title_index:  int

    :param column_index: 根据某列的列索引号进行拆分（从0开始）
    :type column_index: int
    """
    # 将要拆分的工作薄读取为dataframe对象
    # 20210130更新：xlrd库更新到2.x后，无法读取xlsx文件，现在要读取的话需要pd.read_excel(engine='openpyxl')指定引擎
    workbook_df = pd.read_excel(io=workbook_path,sheet_name=sheet_index,header=None,engine='openpyxl')
    # 获取工作薄的表头部分和数据部分，生成2个dataframe
    workbook_header_df = workbook_df[0:title_index]
    workbook_source_df = workbook_df[title_index:]
    # 对单位列去重复,返回一个numpy数组col，包含所有不重复的列字段
    col = workbook_source_df.iloc[:,column_index].unique()
    # 遍历数组
    for x in col:
        # 循环，得到每一个列字段的dataframe
        each_workbook_source_df = workbook_source_df[workbook_source_df.iloc[:,column_index] == x]
        # 拼接dataframe，表头部分与数据部分
        each_workbook_df = workbook_header_df.append(each_workbook_source_df)
        # 将得到的dataframe保存成Excel格式，文件名为列字段。
        each_workbook_df.to_excel(x + '.xlsx', index=False,header=False)

def merge_workbook(filespath:str,sheet_index:int,title_index:int,savename:str) -> None:
    """
    将多个工作薄中同一序号的sheet页合并为1个工作薄
    使用pandas库处理，支持.xlsx类型

    :param filespath: 要合并的表格放在同一个文件夹下，如"D:/xlsx文件夹/"
    :type filespath: str

    :param sheet_index: 表格中的序号（从0开始）
    :type sheet_index： int

    :param title_index: 要合并的表格中表头行数（合并完成后表头只要1个，合并后表格数据由1个表头和多个内容组成）
    :type title_index：int

    :param savename: 保存的文件名称（带路径）
    :type savename：str

    :return: None
    """
    # 获取要合并excel表格的文件名，返回文件夹下所有文件名列表
    fileslist = get_singlefolder_fullfilename(folder_path=filespath,filetype=[".xlsx"])
    # 拿列表中第1个读取后切片，获得表头dataframe
    workbook_header_df = pd.read_excel(io=fileslist[0],sheet_name=sheet_index,header=None,engine='openpyxl')[0:title_index]
    # merge_list 用于循环时添加数据，先将表头dataframe转为列表后赋值
    merge_list = workbook_header_df.values.tolist()
    # 循环对每个文件名通过pd.read_excel读取后，转为列表，extend拼接，不断完善merge_list
    for file in fileslist:
        each_workbook_source_df = pd.read_excel(io=file,sheet_name=sheet_index,header=None)[title_index:]
        each_workbook_source_list = each_workbook_source_df.values.tolist()
        # for l in each_workbook_source_list:
        #     merge_list.append(l)
        merge_list.extend(each_workbook_source_list)
    # merge_list最终转为dataframe，保存为excel文件，即为合并后的文件夹
    workbook_merge_df = DataFrame(merge_list)
    workbook_merge_df.to_excel(excel_writer=savename,index=False,header=False)

def compare_different_dataframe(df1,df2,column_name) -> tuple:
    """
    对比2个pandas DataFrame

    :param df1: 第1个dataframe
    :type df1: pandas.DataFrame

    :param df2: 第2个dataframe
    :type df2: pandas.DataFrame

    :param column_name: 关联的列名
    :type column_name: str

    :return: 包含2个dataframe的元组tuple，tuple[0]是df1有df2没有的行，tuple[1]是df2有df1没有的行
    :rtype: tuple
    """
    # 取df1有而df2没有的
    dfmerge_1 = pd.concat([df1, df2, df2], axis=0)
    dfmerge_1 = dfmerge_1.drop_duplicates(subset=column_name, keep=False)
    # 取df2有但df1没有的
    dfmerge_2 = pd.concat([df2, df1, df1], axis=0)
    dfmerge_2 = dfmerge_2.drop_duplicates(subset=column_name, keep=False)
    return dfmerge_1,dfmerge_2


if __name__ == '__main__':
    split_workbook_by_column(r"D:\MyNutstore\PersonalStudy\PYTHON\WorkOverflow\tests\xlsx\test1.xls",0,1,1)
    # merge_workbook("D:\\MyNutstore\\PersonalStudy\\Python\\帮同事做的小工具\\20201112_谢昊础\\英德市企业缴纳社保相关数据\\",
    #                0,1,"已合并.xlsx")